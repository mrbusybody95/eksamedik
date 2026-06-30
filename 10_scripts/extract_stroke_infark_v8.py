#!/usr/bin/env python3
"""
Ekstraksi Data Stroke Infark — v7
Input : 03_anonymized_text_v6/STROKE_*/ (hasil anonimisasi v6)
Output: 06_extracted_data/stroke_infark_dataset.xlsx

Data per pasien untuk penelitian stroke infark.
Semua field diekstrak dari teks anonim — TANPA PII.

Cara pakai:
  python extract_stroke_infark_v7.py
"""

from pathlib import Path
import re
import json
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

BASE_DIR = Path(__file__).resolve().parent.parent
ANON_DIR = BASE_DIR / "03_anonymized_text_v6"
OUT_DIR = BASE_DIR / "06_extracted_data"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# ============================================================
# UTILITY
# ============================================================

def normalize_text(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text

def clean_snippet(text: str, max_len: int = 500) -> str:
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) > max_len:
        text = text[:max_len] + " ...[truncated]"
    return text

def get_doc_type(filename: str) -> str:
    name = filename.lower()
    if "resume" in name: return "resume"
    if "rad" in name: return "radiology"
    if "lab" in name: return "lab"
    if "cppt_igd" in name: return "cppt_igd"
    if "cppt_ranap" in name or "ranap" in name: return "cppt_ranap"
    return "other"

def read_patient_files(patient_dir: Path):
    files = []
    for fp in sorted(patient_dir.glob("*.anon.txt")):
        text = normalize_text(fp.read_text(encoding="utf-8", errors="ignore"))
        if len(text.strip()) < 20:
            continue
        files.append({
            "source_file": fp.name,
            "doc_type": get_doc_type(fp.name),
            "text": text,
            "char_count": len(text)
        })
    return files

def context(text, start, end, radius=100, max_len=450):
    a = max(0, start - radius)
    b = min(len(text), end + radius)
    return clean_snippet(text[a:b], max_len=max_len)

def first_by_priority(candidates):
    if not candidates: return None
    return sorted(candidates, key=lambda x: (x.get("priority", 9), x.get("pos", 10**9)))[0]

def number_str(x):
    if x is None: return "unknown"
    x = str(x).replace(",", ".")
    m = re.search(r"\d+(?:\.\d+)?", x)
    return m.group(0) if m else "unknown"


# ============================================================
# DEMOGRAFI DARI RESUME
# ============================================================

def extract_demographics(files):
    """Ambil umur, jenis kelamin dari resume."""
    result = {"age": "unknown", "gender": "unknown"}
    for f in files:
        text = f["text"]
        # Umur
        m = re.search(r"(?:Umur|Usia)\s*[:=]?\s*(\d+)\s*(?:Tahun|th|thn)?", text, re.I)
        if m: result["age"] = m.group(1)
        # Jenis Kelamin
        m = re.search(r"(?:JK|Jenis\s*Kelamin|Kelamin)\s*[:=]?\s*(L(?:aki[\s-]*laki)?|P(?:erempuan|ria)?)", text, re.I)
        if m:
            g = m.group(1).upper()
            result["gender"] = "Laki-laki" if g.startswith("L") else "Perempuan"
        if result["age"] != "unknown" and result["gender"] != "unknown":
            break
    return result


# ============================================================
# DIAGNOSIS — KHUSUS STROKE INFARK
# ============================================================

def extract_diagnosis(files):
    """Ekstrak jenis stroke: infark vs hemoragik, plus detail infark."""
    patterns = [
        (r".{0,80}(?:diagnosis|diagnosa|assessment|asesmen|kesimpulan).{0,200}(?:stroke|infark|iskemik|hemoragik|ICH|PIS).{0,200}", 1),
        (r".{0,80}(stroke\s+infark|infark\s+cerebri|stroke\s+iskemik|stroke\s+non\s+hemoragik|SNH|cerebral\s+infarct).{0,150}", 1),
        (r".{0,80}(stroke\s+lakunar|lakunar\s+infark|infark\s+lakuner|small\s+vessel).{0,150}", 2),
        (r".{0,80}(stroke\s+hemoragik|ICH|PIS|perdarahan\s+intraserebral|perdarahan\s+intracerebral).{0,150}", 3),
    ]
    pr = {"resume": 1, "radiology": 2, "cppt_igd": 3, "cppt_ranap": 4}
    candidates = []
    for f in files:
        for pat, _ in patterns:
            for m in re.finditer(pat, f["text"], re.I | re.S):
                candidates.append({
                    "source_file": f["source_file"],
                    "doc_type": f["doc_type"],
                    "priority": pr.get(f["doc_type"], 9),
                    "pos": m.start(),
                    "evidence": clean_snippet(m.group(0), 450)
                })
    best = first_by_priority(candidates)
    if not best:
        return {"stroke_type": "unknown", "diagnosis_text": "unknown", "evidence": ""}

    low = best["evidence"].lower()
    infark_terms = ["infark", "infarct", "iskemik", "ischemic", "lakunar", "lacunar", "snh", "non hemoragik"]
    hemoragik_terms = ["hemoragik", "hemorrhagic", "ich", "pis", "perdarahan intraserebral", "perdarahan intracerebral"]
    
    is_infark = any(t in low for t in infark_terms)
    is_hemo = any(t in low for t in hemoragik_terms)

    if is_infark and not is_hemo: stype = "INFARK"
    elif is_hemo and not is_infark: stype = "HEMORAGIK"
    elif is_infark and is_hemo: stype = "MIXED"
    else: stype = "UNCLEAR"

    return {
        "stroke_type": stype,
        "diagnosis_text": best["evidence"],
        "evidence": best["evidence"]
    }


# ============================================================
# GCS, VITAL SIGN
# ============================================================

def extract_gcs(files):
    candidates = []
    pr = {"cppt_igd": 1, "resume": 2, "cppt_ranap": 3}
    
    component_patterns = [
        r"\bGCS\s*[:=]?\s*E\s*(\d)\s*M\s*(\d)\s*V\s*(\d)\b",
        r"\bE\s*(\d)\s*M\s*(\d)\s*V\s*(\d)\b",
    ]
    total_patterns = [
        r"\bGCS\s*[:=]?\s*(1[0-5]|[3-9])\b",
    ]
    
    for f in files:
        if f["doc_type"] not in pr: continue
        for pat in total_patterns:
            for m in re.finditer(pat, f["text"], re.I):
                candidates.append({"value": m.group(1), "source_file": f["source_file"],
                    "priority": pr[f["doc_type"]], "pos": m.start(),
                    "evidence": context(f["text"], m.start(), m.end())})
        for pat in component_patterns:
            for m in re.finditer(pat, f["text"], re.I):
                nums = [int(x) for x in m.groups()]
                candidates.append({"value": str(sum(nums)), "source_file": f["source_file"],
                    "priority": pr[f["doc_type"]], "pos": m.start(),
                    "evidence": context(f["text"], m.start(), m.end())})
    
    best = first_by_priority(candidates)
    return best["value"] if best else "unknown"


def extract_vitals(files):
    """TD, HR, RR, Suhu, SpO2 — nilai awal dari CPPT IGD."""
    result = {"td_sistol": "unknown", "td_diastol": "unknown", "hr": "unknown",
              "rr": "unknown", "suhu": "unknown", "spo2": "unknown"}
    pr = {"cppt_igd": 1, "resume": 2, "cppt_ranap": 3}
    
    patterns = {
        "td": (r"\bTD\s*[:=]?\s*(\d{2,3})\s*/\s*(\d{2,3})\b", lambda m: (m.group(1), m.group(2))),
        "hr": (r"\b(?:HR|Nadi|Heart\s*Rate)\s*[:=]?\s*(\d{2,3})\b", lambda m: m.group(1)),
        "rr": (r"\b(?:RR|Respiratory|Resp)\s*[:=]?\s*(\d{1,3})\b", lambda m: m.group(1)),
        "suhu": (r"\b[SST]\s*[:=]?\s*(\d{2,3}(?:[.,]\d)?)\s*(?:[°′]?C)?\b", lambda m: m.group(1).replace(",", ".")),
        "spo2": (r"\b(?:SPO2|SpO2|SaO2|O2\s*Sat)\s*[:=]?\s*(\d{2,3})\s*%?\b", lambda m: m.group(1)),
    }
    
    for f in files:
        if f["doc_type"] not in pr: continue
        text = f["text"]
        for key, (pat, extractor) in patterns.items():
            if key == "td":
                if result["td_sistol"] != "unknown": continue
                for m in re.finditer(pat, text, re.I):
                    if 50 <= int(m.group(1)) <= 300 and 30 <= int(m.group(2)) <= 200:
                        result["td_sistol"], result["td_diastol"] = m.group(1), m.group(2)
                        break
            else:
                if result[key] != "unknown": continue
                m = re.search(pat, text, re.I)
                if m:
                    result[key] = extractor(m)
    
    return result


# ============================================================
# CT SCAN — LOKASI INFARK
# ============================================================

def extract_ct_scan(files):
    """Ekstrak hasil CT Scan kepala — khusus lokasi infark."""
    result = {
        "ct_documented": "no", "ct_result": "", "ct_infark_lokasi": "",
        "ct_aspects": "unknown", "ct_perdarahan": "unknown",
        "ct_midline_shift": "unknown", "ct_hidrosefalus": "unknown",
        "ct_atrofi": "unknown"
    }
    
    for f in files:
        if f["doc_type"] not in ["radiology", "resume"]: continue
        text = f["text"]
        low = text.lower()
        
        # Cek apakah ada CT scan kepala
        if re.search(r"(?:ct\s*scan|msct|ct\s*kepala|ct\s*head|head\s*ct)", low):
            result["ct_documented"] = "yes"
        
        # Cari lokasi infark
        infark_locs = set()
        loc_patterns = [
            "thalamus", "talamus", "kapsula\s*interna", "capsula\s*interna",
            "nukleus\s*lentiformis", "nucleus\s*lentiformis", "ganglia\s*basalis",
            "basal\s*ganglia", "pons", "cerebellum", "serebelum",
            "centrum\s*semiovale", "corona\s*radiata", "periventrikel",
            "frontal", "parietal", "temporal", "occipital", "oksipital",
            "subcortical", "subkortikal", "lacunar", "lakunar",
            "mca", "aca", "pca", "sistem\s*karotis", "vertebrobasiler"
        ]
        for pat in loc_patterns:
            if re.search(pat, low):
                # Ambil konteks untuk konfirmasi infark
                m = re.search(r".{0,40}" + pat + r".{0,60}", low, re.I)
                if m and ("infark" in m.group() or "infarct" in m.group() or "lakuner" in m.group() or "hipodens" in m.group()):
                    infark_locs.add(pat.replace("\\s*", " ").replace("\\s?", ""))
        
        if infark_locs:
            result["ct_infark_lokasi"] = ", ".join(sorted(infark_locs))
        
        # ASPECTS
        m = re.search(r"\bASPECTS\s*[:=]?\s*(\d{1,2})\b", text, re.I)
        if m: result["ct_aspects"] = m.group(1)
        
        # Perdarahan
        if re.search(r"(?:tidak\s+tampak\s+tanda[-\s]tanda\s+perdarahan|tidak\s+tampak\s+perdarahan|no\s+hemorrhage|no\s+bleeding)", low):
            result["ct_perdarahan"] = "tidak_ada"
        elif re.search(r"\b(?:perdarahan|hemorrhage|ICH|PIS|SAH|SAB|IVH|intraserebral|intracerebral|subarachnoid|intraventrikel)", low):
            result["ct_perdarahan"] = "ada"
        
        # Midline shift
        if re.search(r"(?:midline\s*shift|pergeseran\s*garis\s*tengah|mid\s*line\s*shift)", low):
            if re.search(r"(?:tidak\s+tampak|no|tanpa)", low[:low.find("midline")][-40:]) if "midline" in low else False:
                result["ct_midline_shift"] = "tidak_ada"
            else:
                result["ct_midline_shift"] = "ada"
        
        # Hidrosefalus
        if re.search(r"(?:hidrosefalus|hydrocephalus)", low):
            result["ct_hidrosefalus"] = "ada"
        
        # Atrofi
        if re.search(r"(?:atrofi|atrophy)", low):
            result["ct_atrofi"] = "ada"
        
        # Ambil teks hasil CT
        for pat in [r"(?:KESIMPULAN|Kesimpulan|kesimpulan)\s*[:=]?\s*(.{20,800})",
                     r"(?:IMPRESSION|Impression)\s*[:=]?\s*(.{20,800})"]:
            m = re.search(pat, text, re.I | re.S)
            if m:
                result["ct_result"] = clean_snippet(m.group(1), 600)
                break
    
    return result


# ============================================================
# LAB
# ============================================================

LAB_PATTERNS = {
    "hb": [r"\bHb\s*[:=]?\s*(\d{1,2}[,.]?\d*)", r"\bHemoglobin\s*[:=]?\s*(\d{1,2}[,.]?\d*)"],
    "leukosit": [r"\b(?:Leukosit|Lekosit)\s*[:=]?\s*(\d{1,3}[,.]?\d*)", r"\bWBC\s*[:=]?\s*(\d{1,3}[,.]?\d*)"],
    "trombosit": [r"\b(?:Trombosit|Trombosit|Platelet)\s*[:=]?\s*(\d{1,4}[,.]?\d*)", r"\bPLT\s*[:=]?\s*(\d{1,4}[,.]?\d*)"],
    "ht": [r"\b(?:Ht|Hematokrit|Hematocrit)\s*[:=]?\s*(\d{1,3}[,.]?\d*)"],
    "eritrosit": [r"\b(?:Eritrosit|Erythrocyte|Red\s*Blood\s*Cell)\s*[:=]?\s*(\d{1,3}[,.]?\d*)"],
    "gds": [r"\bGDS\s*[:=]?\s*(\d{2,4}[,.]?\d*)", r"\bGlukosa\s*Darah\s*Sewaktu\s*[:=]?\s*(\d{2,4}[,.]?\d*)"],
    "ureum": [r"\bUreum\s*[:=]?\s*(\d{1,4}[,.]?\d*)"],
    "kreatinin": [r"\b(?:Kreatinin|Creatinine)\s*[:=]?\s*(\d{1,3}[,.]?\d*)", r"\bCr\s*[:=]?\s*(\d{1,3}[,.]?\d*)"],
    "natrium": [r"\bNatrium\s*\(Na\)\s*-?\s*(\d{2,3}[,.]?\d*)", r"\bNatrium\s*[:=]?\s*(\d{2,3}[,.]?\d*)", r"\bSodium\s*[:=]?\s*(\d{2,3}[,.]?\d*)"],
    "kalium": [r"\bKalium\s*\(K\)\s*-?\s*(\d{1,2}[,.]?\d*)", r"\bKalium\s*[:=]?\s*(\d{1,2}[,.]?\d*)", r"\bPotassium\s*[:=]?\s*(\d{1,2}[,.]?\d*)"],
    "asam_urat": [r"\b(?:Asam\s*Urat|Uric\s*Acid)\s*[:=]?\s*(\d{1,3}[,.]?\d*)"],
    "inr": [r"\bINR\s*[:=]?\s*(\d{1,3}[,.]?\d*)"],
    "ldl": [r"\bLDL\s*[:=]?\s*(\d{1,4}[,.]?\d*)"],
    "kolesterol_total": [r"\b(?:Kolesterol\s*Total|Cholesterol\s*Total)\s*[:=]?\s*(\d{1,4}[,.]?\d*)"],
    "trigliserida": [r"\b(?:Trigliserida|Triglyceride)\s*[:=]?\s*(\d{1,4}[,.]?\d*)"],
    "hba1c": [r"\bHbA1c\s*[:=]?\s*(\d{1,2}[,.]?\d*)"],
}

def extract_labs(files):
    result = {k: "unknown" for k in LAB_PATTERNS}
    pr = {"lab": 1, "resume": 2, "cppt_ranap": 3, "cppt_igd": 4}
    
    for lab_key, patterns in LAB_PATTERNS.items():
        candidates = []
        for f in files:
            if f["doc_type"] not in pr: continue
            for pat in patterns:
                for m in re.finditer(pat, f["text"], re.I):
                    candidates.append({
                        "value": number_str(m.group(1)),
                        "priority": pr[f["doc_type"]],
                        "pos": m.start()
                    })
        best = first_by_priority(candidates)
        if best: result[lab_key] = best["value"]
    
    return result


# ============================================================
# TANGGAL PEMERIKSAAN LAB & RADIOLOGI
# ============================================================

def parse_date_indonesian(text: str) -> str | None:
    """Parse Indonesian date string like '03 Mei 2026' → '2026-05-03'."""
    bulan = {
        "januari":"01", "februari":"02", "maret":"03", "april":"04",
        "mei":"05", "juni":"06", "juli":"07", "agustus":"08",
        "september":"09", "oktober":"10", "november":"11", "desember":"12",
        "january":"01", "february":"02", "march":"03", "april":"04",
        "may":"05", "june":"06", "july":"07", "august":"08",
    }
    m = re.search(r"(\d{1,2})\s+(Mei|Januari|Februari|Maret|April|Juni|Juli|Agustus|September|Oktober|November|Desember|January|February|March|April|May|June|July|August)\s+(\d{4})", text, re.I)
    if m:
        bln = bulan.get(m.group(2).lower(), "??")
        return f"{m.group(3)}-{bln}-{int(m.group(1)):02d}"
    return None

def extract_lab_dates(files):
    """Ambil tanggal sampling lab & tanggal radiologi (CT & thorax)."""
    result = {
        "lab_tanggal_igd": "unknown",
        "lab_tanggal_ranap": "unknown",
        "lab_tanggal_pertama": "unknown",
        "ct_tanggal": "unknown",
        "thorax_tanggal": "unknown",
    }
    
    for f in files:
        text = f["text"]
        
        if f["doc_type"] == "lab":
            # Tgl. Sampling: 03 Mei 2026 07:17:
            m = re.search(r"Tgl\.?\s*Sampling\s*[:=]?\s*(\d{1,2}\s+\w+\s+\d{4})", text, re.I)
            if m:
                parsed = parse_date_indonesian(m.group(1))
                if parsed:
                    # Determine IGD vs rawat inap
                    if "igd" in f["source_file"].lower():
                        result["lab_tanggal_igd"] = parsed
                    else:
                        result["lab_tanggal_ranap"] = parsed
                    if result["lab_tanggal_pertama"] == "unknown":
                        result["lab_tanggal_pertama"] = parsed
                    else:
                        # Keep earliest
                        if parsed < result["lab_tanggal_pertama"]:
                            result["lab_tanggal_pertama"] = parsed
        
        if f["doc_type"] == "radiology":
            text_upper = text.upper()
            low = text.lower()
            # TANGGAL SELESAI : 01 MEI 2026
            m = re.search(r"TANGGAL\s*(?:SELESAI|PEMERIKSAAN|SELESEI|SELESED|SELES)\s*[:=]?\s*(\d{1,2}\s+\w+\s+\d{4})", text_upper)
            if not m:
                m = re.search(r"Tgl\.?\s*Pmk\s*[:=]?\s*(\d{1,2}\s+\w+\s+\d{4})", text, re.I)
            if m:
                parsed = parse_date_indonesian(m.group(1))
                if parsed:
                    # Bisa jadi 1 file berisi CT + Thorax (combined report)
                    is_ct = bool(re.search(r"(?:ct\s*scan|msct|ct\s*kepala|ct\s*head|head\s*ct)", low))
                    is_thorax = bool(re.search(r"(?:thorax|foto\s*polos\s*dada|chest\s*x.ray|x.ray\s*thorax|foto\s*thorax)\s", low))
                    
                    if is_thorax and result["thorax_tanggal"] == "unknown":
                        result["thorax_tanggal"] = parsed
                    if is_ct and result["ct_tanggal"] == "unknown":
                        result["ct_tanggal"] = parsed
    
    return result


# ============================================================
# THORAX FOTO — HASIL FOTO POLOS DADA
# ============================================================

def extract_thorax(files):
    """Ekstrak hasil foto thorax dari resume & radiology."""
    result = {
        "thorax_documented": "tidak",
        "thorax_kesan": "unknown",
    }
    
    for f in files:
        if f["doc_type"] not in ["radiology", "resume"]: continue
        text = f["text"]
        low = text.lower()
        
        # Deteksi apakah ada thorax
        is_thorax = bool(re.search(r"(?:thorax|foto\s*polos\s*dada|chest\s*x.ray|x.ray\s*thorax|foto\s*thorax)", low))
        if not is_thorax:
            continue
        
        result["thorax_documented"] = "ada"
        
        # Ambil KESAN dari thorax
        # Cari KESAN yang mengandung keyword thorax
        m = re.search(r"KESAN\s*:\s*(.{20,300})", text, re.I | re.S)
        if m:
            kesan = m.group(1).strip()
            # Filter: hanya ambil jika konteks thorax
            if re.search(r"(?:broncho|pneumonia|cardiomegali|paru|jantung|aorta|infiltrat|efusi)", kesan, re.I):
                result["thorax_kesan"] = clean_snippet(kesan, 300)
    
    return result


# ============================================================
# DEMAM
# ============================================================

def extract_demam(files):
    """Deteksi apakah pasien mengalami demam saat masuk & selama rawat."""
    result = {
        "demam_saat_masuk": "tidak",
        "demam_selama_rawat": "tidak",
        "suhu_tertinggi": "unknown",
    }
    
    demam_keywords = [r"\bfebris\b", r"\bdemam\b", r"\bpanas\s*badan\b", r"\bpireksia\b", r"\bhiperpireksia\b"]
    suhu_max = 0.0
    
    for f in files:
        text = f["text"]
        low = text.lower()
        
        # Saat masuk (IGD)
        if f["doc_type"] == "cppt_igd":
            for pat in demam_keywords:
                if re.search(pat, low):
                    result["demam_saat_masuk"] = "ya"
                    break
        
        # Selama rawat (CPPT Ranap)
        if f["doc_type"] == "cppt_ranap":
            for pat in demam_keywords:
                if re.search(pat, low):
                    result["demam_selama_rawat"] = "ya"
                    break
        
        # Suhu tertinggi dari semua dokumen
        for m in re.finditer(r"\bS(?:uhu)?\s*[:=]?\s*(\d{2,3}[.,]\d)\s*(?:[°′]?C)?", text, re.I):
            try:
                suhu = float(m.group(1).replace(",", "."))
                if 35 <= suhu <= 42 and suhu > suhu_max:
                    suhu_max = suhu
            except ValueError:
                pass
    
    if suhu_max > 37.5:
        result["suhu_tertinggi"] = str(suhu_max)
    else:
        result["suhu_tertinggi"] = "unknown"
    
    # Jika demam_saat_masuk masih tidak tapi suhu tinggi...
    if result["demam_saat_masuk"] == "tidak" and result["suhu_tertinggi"] != "unknown":
        if float(result["suhu_tertinggi"]) >= 38:
            result["demam_saat_masuk"] = "ya"
    
    return result


# ============================================================
# FAKTOR RISIKO & KOMORBID
# ============================================================

def extract_risk_factors(files):
    """Cari faktor risiko stroke dari semua dokumen."""
    result = {
        "hipertensi": "unknown",
        "diabetes_melitus": "unknown",
        "dislipidemia": "unknown",
        "atrial_fibrilasi": "unknown",
        "penyakit_jantung": "unknown",
        "stroke_sebelumnya": "unknown",
        "merokok": "unknown",
        "obesitas": "unknown",
    }
    
    # Keywords per faktor risiko
    risk_map = {
        "hipertensi": [r"\b(?:hipertensi|hypertension|HT|HD|darah\s*tinggi)\b"],
        "diabetes_melitus": [r"\b(?:diabetes?|DM|diabetes\s*melitus|kencing\s*manis|gula\s*darah|hiperglikemi)\b"],
        "dislipidemia": [r"\b(?:dislipidemia|dyslipidemia|kolesterol|hiperlipidemia|hiperkolesterolemia)\b"],
        "atrial_fibrilasi": [r"\b(?:atrial\s*fibrilasi|atrial\s*fibrillation|AF|AFib|CAF|flutter)\b"],
        "penyakit_jantung": [r"\b(?:jantung|cardio|CAD|PJK|kardiomegali|CHF|gagal\s*jantung|heart\s*failure|IHD)\b"],
        "stroke_sebelumnya": [r"\b(?:stroke\s*ulang|riwayat\s*stroke|stroke\s*sebelumnya|previous\s*stroke|stroke\s*ec\s+infark)\b"],
        "merokok": [r"\b(?:merokok|smoking|perokok|rokok|smoker)\b"],
        "obesitas": [r"\b(?:obesitas|obesity|obes|IMT\s*>=?3[05]|BMI\s*>=?3[05])\b"],
    }
    
    for f in files:
        text = f["text"].lower()
        for risk, patterns in risk_map.items():
            if result[risk] != "unknown": continue
            for pat in patterns:
                if re.search(pat, text):
                    result[risk] = "ada"
                    break
    
    # Default unknown → "tidak_tercatat"
    for k in result:
        if result[k] == "unknown":
            result[k] = "tidak_tercatat"
    
    return result


# ============================================================
# RIWAYAT PENYAKIT TERDAHULU (RPD)
# ============================================================

def extract_rpd(files):
    """Ekstrak teks Riwayat Penyakit Terdahulu."""
    for f in files:
        if f["doc_type"] in ["resume", "cppt_igd"]:
            m = re.search(r"(?:RPD|Riwayat\s*Penyakit\s*(?:Terdahulu|Dahulu|Sebelumnya))\s*[:=]?\s*(.{20,300})", f["text"], re.I)
            if m:
                return clean_snippet(m.group(1), 300)
    return "unknown"


# ============================================================
# OBAT-OBATAN STROKE INFARK
# ============================================================

def extract_medications(files):
    """Deteksi obat-obatan yang diberikan — khusus stroke infark."""
    result = {}
    
    med_map = {
        "antiplatelet_aspirin": [r"\baspirin\b", r"\basetosal\b", r"\bascardia\b"],
        "antiplatelet_clopidogrel": [r"\bclopidogrel\b", r"\bclopid\b"],
        "antiplatelet_cilostazol": [r"\bcilostazol\b"],
        "antiplatelet_ticagrelor": [r"\bticagrelor\b"],
        "antikoagulan": [r"\bwarfarin\b", r"\bheparin\b", r"\benoxaparin\b", r"\brivaroxaban\b", r"\bapixaban\b", r"\bdabigatran\b"],
        "statin": [r"\bstatin\b", r"\batorvastatin\b", r"\bsimvastatin\b", r"\brosuvastatin\b"],
        "antihipertensi": [r"\bamlodipin", r"\bcaptopril\b", r"\bnicardipin", r"\blabetalol\b", r"\bvalsartan\b", r"\bcandesartan\b", r"\bbisoprolol\b"],
        "mannitol": [r"\b(?:manitol|mannitol)\b"],
        "citicoline": [r"\bciticoline\b", r"\bsitikolin\b", r"\bciticholine\b"],
        "ppi": [r"\bomeprazole\b", r"\bpantoprazole\b", r"\brabeprazole\b", r"\blansoprazole\b", r"\bsucralfat\b"],
        "antibiotik": [r"\bceftriaxone\b", r"\blevofloxacin\b", r"\bmoxifloxacin\b", r"\bazithromycin\b", r"\bcefixime\b", r"\bcefotaxime\b", r"\bmetronidazole\b"],
        "antiepilepsi": [r"\blevetiracetam\b", r"\bphenytoin\b", r"\bfenitoin\b", r"\bvalproat\b"],
        "insulin": [r"\binsulin\b", r"\bnovorapid\b", r"\blantus\b"],
    }
    
    for f in files:
        text = f["text"].lower()
        for key, patterns in med_map.items():
            if key not in result or result.get(key) == "tidak_tercatat":
                for pat in patterns:
                    m = re.search(pat, text)
                    if m:
                        result[key] = "diberikan"
                        break
    
    # Default
    for key in med_map:
        if key not in result:
            result[key] = "tidak_tercatat"
    
    return result


# ============================================================
# TINDAKAN & KONSULTASI
# ============================================================

def extract_actions(files):
    """Deteksi tindakan/konsultasi yang dilakukan."""
    result = {}
    
    action_map = {
        "konsul_neurologi": [r"\bkonsul\s*neuro", r"\bsp\s*[ns]\b", r"\bdokter\s*saraf\b"],
        "konsul_bedah_saraf": [r"\bbedah\s*saraf\b", r"\bsp\s*bs\b"],
        "konsul_jantung": [r"\bkonsul\s*jantung", r"\bsp\s*jp\b", r"\bekg\b"],
        "rawat_icu_hcu": [r"\b(?:ICU|HCU|intensive\s*care|NICU|PICU)\b"],
        "fisioterapi": [r"\bfisioterapi\b", r"\brehab\b", r"\bmobilisasi\b"],
        "konsul_gizi": [r"\bkonsul\s*gizi\b", r"\b(?:ahli\s*)?gizi\b"],
        "skrining_menelan": [r"\b(?:skrining\s*menelan|swallowing|disfagia|tes\s*menelan)\b"],
        "edukasi_keluarga": [r"\bedukasi\b", r"\binformed\s*consent\b", r"\bpenjelasan\s*keluarga\b"],
    }
    
    for f in files:
        text = f["text"].lower()
        for key, patterns in action_map.items():
            if key not in result or result.get(key) == "tidak_tercatat":
                for pat in patterns:
                    if re.search(pat, text):
                        result[key] = "dilakukan"
                        break
    
    for key in action_map:
        if key not in result:
            result[key] = "tidak_tercatat"
    
    return result


# ============================================================
# OUTCOME
# ============================================================

def extract_outcome(files):
    """Cara keluar, lama rawat, kondisi pulang."""
    result = {
        "cara_keluar": "unknown",
        "lama_rawat_hari": "unknown",
        "kondisi_pulang": "unknown",
        "rencana_kontrol": "unknown"
    }

    tgl_masuk = None
    tgl_keluar = None

    for f in files:
        if f["doc_type"] not in ["resume"]: continue
        text = f["text"]

        # Tanggal Masuk
        m = re.search(r"Tanggal\s*Masuk\s*[:=]?\s*(\d{1,2}\s+\w+\s+\d{4})", text, re.I)
        if m:
            parsed = parse_date_indonesian(m.group(1))
            if parsed: tgl_masuk = parsed

        # Tanggal Keluar
        m = re.search(r"Tanggal\s*Keluar\s*[:=]?\s*(\d{1,2}\s+\w+\s+\d{4})", text, re.I)
        if m:
            parsed = parse_date_indonesian(m.group(1))
            if parsed: tgl_keluar = parsed

        # Cara keluar
        m = re.search(r"(?:Cara\s*(?:Pasien\s*)?Keluar|Discharge)\s*[:=]?\s*(.{10,100})", text, re.I)
        if m:
            v = m.group(1).strip()
            if re.search(r"(?:diizinkan|pulang|boleh)", v, re.I): result["cara_keluar"] = "pulang"
            elif re.search(r"(?:APS|atas\s*permintaan\s*sendiri)", v, re.I): result["cara_keluar"] = "APS"
            elif re.search(r"(?:meninggal|meninggal dunia|death)", v, re.I): result["cara_keluar"] = "meninggal"
            elif re.search(r"(?:rujuk|dirujuk)", v, re.I): result["cara_keluar"] = "dirujuk"
            else: result["cara_keluar"] = clean_snippet(v, 80)

        # Lama rawat (direct mention)
        m = re.search(r"(?:Lama\s*Rawat|Length\s*of\s*Stay|LOS|Lama\s*Dirawat)\s*[:=]?\s*(\d+)\s*(?:hari|hr|day)?", text, re.I)
        if m: result["lama_rawat_hari"] = m.group(1)

        # Kondisi pulang
        m = re.search(r"(?:Kondisi\s*Pasien\s*(?:Saat\s*)?(?:Pulang|Keluar)|Condition\s*at\s*Discharge)\s*[:=]?\s*(.{10,100})", text, re.I)
        if m:
            v = m.group(1).strip()
            if re.search(r"(?:perbaikan|membaik|baik|improved|stable)", v, re.I): result["kondisi_pulang"] = "perbaikan"
            elif re.search(r"(?:tetap|sama|unchanged)", v, re.I): result["kondisi_pulang"] = "tetap"
            elif re.search(r"(?:memburuk|worsen)", v, re.I): result["kondisi_pulang"] = "memburuk"
            else: result["kondisi_pulang"] = clean_snippet(v, 80)

        # Rencana kontrol
        m = re.search(r"(?:Kontrol|Follow.?up|Control)\s*[:=]?\s*(.{10,100})", text, re.I)
        if m: result["rencana_kontrol"] = clean_snippet(m.group(1), 100)

    # Hitung lama rawat dari tanggal jika belum ketemu
    if result["lama_rawat_hari"] == "unknown" and tgl_masuk and tgl_keluar:
        try:
            from datetime import datetime
            d1 = datetime.strptime(tgl_masuk, "%Y-%m-%d")
            d2 = datetime.strptime(tgl_keluar, "%Y-%m-%d")
            delta = (d2 - d1).days
            if delta >= 0:
                result["lama_rawat_hari"] = str(delta)
        except ValueError:
            pass

    return result


# ============================================================
# MAIN: PROSES SEMUA PASIEN
# ============================================================

def process_all_patients():
    patient_dirs = sorted([d for d in ANON_DIR.iterdir() if d.is_dir() and d.name.startswith("STROKE_")])
    print(f"Memproses {len(patient_dirs)} pasien dari {ANON_DIR}...")
    
    dataset = []
    
    for pdir in patient_dirs:
        files = read_patient_files(pdir)
        if not files:
            print(f"  ⚠️  {pdir.name}: tidak ada file (skip)")
            continue
        
        print(f"  📄 {pdir.name}: {len(files)} file...")
        
        row = {"patient_id": pdir.name}
        
        # Demografi
        demo = extract_demographics(files)
        row.update({f"demo_{k}": v for k, v in demo.items()})
        
        # Diagnosis
        diag = extract_diagnosis(files)
        row["stroke_type"] = diag["stroke_type"]
        row["diagnosis_text"] = diag["diagnosis_text"]
        
        # GCS
        row["gcs_initial"] = extract_gcs(files)
        
        # Vital signs
        vitals = extract_vitals(files)
        row.update({f"vital_{k}": v for k, v in vitals.items()})
        
        # CT Scan
        ct = extract_ct_scan(files)
        row.update({f"ct_{k}": v for k, v in ct.items()})
        
        # Lab
        labs = extract_labs(files)
        row.update({f"lab_{k}": v for k, v in labs.items()})

        # Tanggal lab & radiologi
        lab_dates = extract_lab_dates(files)
        row.update(lab_dates)

        # Thorax
        thorax = extract_thorax(files)
        row.update({f"thorax_{k}": v for k, v in thorax.items()})

        # Demam
        demam = extract_demam(files)
        row.update(demam)

        # Risk factors
        risks = extract_risk_factors(files)
        row.update({f"rf_{k}": v for k, v in risks.items()})
        
        # RPD
        row["rpd_text"] = extract_rpd(files)
        
        # Medications
        meds = extract_medications(files)
        row.update({f"med_{k}": v for k, v in meds.items()})
        
        # Actions
        actions = extract_actions(files)
        row.update({f"act_{k}": v for k, v in actions.items()})
        
        # Outcome
        outcome = extract_outcome(files)
        row.update({f"outcome_{k}": v for k, v in outcome.items()})
        
        dataset.append(row)
        print(f"    ✅ {row.get('stroke_type','?')} | GCS={row.get('gcs_initial','?')} | "
              f"TD {row.get('vital_td_sistol','?')}/{row.get('vital_td_diastol','?')} | "
              f"Demam={row.get('demam_saat_masuk','?')} | "
              f"Lab={row.get('lab_tanggal_pertama','?')} | "
              f"Lama={row.get('outcome_lama_rawat_hari','?')} hari")
    
    return dataset


def save_to_excel(dataset):
    if not dataset:
        print("Tidak ada data untuk disimpan.")
        return
    
    xlsx_path = OUT_DIR / "stroke_infark_dataset.xlsx"
    csv_path = OUT_DIR / "stroke_infark_dataset.csv"
    
    keys = list(dataset[0].keys())
    
    # CSV
    import csv
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=keys)
        w.writeheader()
        w.writerows(dataset)
    print(f"\n📊 CSV: {csv_path} ({len(dataset)} baris)")
    
    # Excel
    if OPENPYXL_OK:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stroke Infark Data"
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Write header
        for col, key in enumerate(keys, 1):
            cell = ws.cell(row=1, column=col, value=key)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        
        # Write data
        for row_idx, row_data in enumerate(dataset, 2):
            for col_idx, key in enumerate(keys, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, ""))
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Column widths
        col_widths = {
            "patient_id": 20, "stroke_type": 12, "gcs_initial": 10,
            "diagnosis_text": 60, "rpd_text": 40,
            "thorax_kesan": 50, "ct_ct_result": 50,
            "lab_tanggal_pertama": 16, "lab_tanggal_igd": 16,
            "lab_tanggal_ranap": 16, "ct_tanggal": 16, "thorax_tanggal": 16,
        }
        for col, key in enumerate(keys, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = col_widths.get(key, 15)
        
        # Freeze panes
        ws.freeze_panes = "B2"
        
        # Auto filter
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(keys))}{len(dataset)+1}"
        
        wb.save(xlsx_path)
        print(f"📊 Excel: {xlsx_path}")
    else:
        print("⚠️  openpyxl tidak tersedia. Hanya CSV yang disimpan.")
    
    # Summary
    print("\n" + "="*60)
    print("RINGKASAN DATASET")
    print("="*60)
    for row in dataset:
        print(f"  {row['patient_id']:25s} | {row['stroke_type']:10s} | "
              f"GCS={row['gcs_initial']:2s} | TD {row['vital_td_sistol']}/{row['vital_td_diastol']}")
    
    infark_count = sum(1 for r in dataset if r.get("stroke_type") == "INFARK")
    print(f"\nTotal pasien: {len(dataset)}")
    print(f"Stroke Infark: {infark_count}")
    print(f"Stroke Hemoragik: {sum(1 for r in dataset if r.get('stroke_type') == 'HEMORAGIK')}")
    print(f"Unknown: {sum(1 for r in dataset if r.get('stroke_type') not in ('INFARK', 'HEMORAGIK'))}")


if __name__ == "__main__":
    t0 = datetime.now()
    dataset = process_all_patients()
    save_to_excel(dataset)
    elapsed = (datetime.now() - t0).total_seconds()
    print(f"\n⏱  Selesai dalam {elapsed:.1f} detik")
